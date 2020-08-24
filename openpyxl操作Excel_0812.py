# -*- coding: utf-8 -*-
"""
Created on Thu Aug 13 09:16:39 2020

@author: lvfanghu
"""

from openpyxl import Workbook

#创建一个空的工作簿
wb = Workbook()
#保存文件
wb.save('E:\\study\\全栈数据分析\\data/wb0812.xlsx')

#创建一个sheet，指定主题为“ws”，指定位置为0（第一个）
ws = wb.create_sheet('ws',0)
#不指定位置会在后面追加一个sheet
ws2 = wb.create_sheet('ws2')
#-1表示倒数第二个
ws3 = wb.create_sheet('ws3',-1)
wb.save('E:\\study\\全栈数据分析\\data/wb0812.xlsx')


#读取文件
from openpyxl import load_workbook

wb = load_workbook('E:\\study\\全栈数据分析\\data/wb0812.xlsx')

#获取所有sheet的名字
print(wb.sheetnames)
#根据sheet名获取sheet
sheet = wb['Sheet']
#修改sheet的主题
sheet.title = '表一'
print(wb.sheetnames)
wb.save('E:\\study\\全栈数据分析\\data/wb0812.xlsx')

#读取表中数据
wb = load_workbook('E:\\study\\全栈数据分析\\data/读写.xlsx')
#获取表
sheet1 = wb['读']
#获取行列数
row = sheet1.max_row
column = sheet1.max_column
#获取一列
one_column = sheet1['A']
print('一列：',one_column)
#获取一行
one_row = sheet1['1']
print('一行：',one_row)
#获取一个单元格
one_cell = sheet1['A1']
print('一个单元格：',one_cell)
print('A1的值：',one_cell.value)

#向表中写入数据
wb = load_workbook('E:\\study\\全栈数据分析\\data/读写.xlsx')
#获取表
sheet = wb['写']
#表中单元格写入数据
sheet['A1'].value = 1
#写入一列
#法一
for i in range(1,11):
    sheet['A{}'.format(i)].value = i
# =============================================================================
# #法二
# for i in range(1,11):
#     sheet.cell(column=1,row=i,value=i)
# wb.save('E:\\study\\全栈数据分析\\data/读写.xlsx')
# =============================================================================

#写入一行
for j in range(1,11):
    sheet.cell(row=11,column=j,value=j)

wb.save('E:\\study\\全栈数据分析\\data/读写.xlsx')

#添加计算列数据
wb = load_workbook('E:\\study\\全栈数据分析\\0812\\excle/案例.xlsx')
sheet = wb['订单']

print('行数：',sheet.max_row)
print('列数：',sheet.max_column)

#获取单价和数量列，增加销售额列
sheet.cell(row=1,column=19,value='销售额')
#获取列数据
for row_index in range(2,sheet.max_row+1):
    #单价数据
    one_price = sheet.cell(row = row_index,column = 17).value
    #数量
    one_num = sheet.cell(row = row_index,column = 18).value
    
    if one_price == None or one_num == None:
        sheet.cell(row = row_index,column = 19).value = None
    else:
        sheet.cell(row = row_index,column =19).value = one_price * one_num
        
wb.save('E:\\study\\全栈数据分析\\0812\\excle/案例.xlsx')

#根据各个省的订单量绘制柱状图
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import BarChart,Reference

wb = load_workbook('E:\\study\\全栈数据分析\\0812\\excle/案例.xlsx')
sheet = wb['各省销量']

#创建柱状图对象
bar_chart = BarChart()

labels = Reference(sheet,min_row = 2,max_row = 32,min_col = 1,max_col = 1)
data = Reference(sheet,min_row = 2,max_row = 32,min_col = 2,max_col = 2)

bar_chart.add_data(data)
bar_chart.set_categories(labels)#设置图例
sheet.add_chart(bar_chart,'D10')

wb.save('E:\\study\\全栈数据分析\\0812\\excle/案例.xlsx')

#将本地图片添加到指定位置
from openpyxl.drawing.image import Image
wb = load_workbook('E:\\study\\全栈数据分析\\0812\\excle/案例.xlsx')
wb.create_sheet('image')

sheet = wb['image']
img = Image('D:/BaiduNetdiskDownload/壁纸/1418704224888.jpeg')
sheet.add_image(img,'B2')
wb.save('E:\\study\\全栈数据分析\\0812\\excle/案例.xlsx')

#修改单元格样式
from openpyxl.styles import Alignment,Border,Side,Font
from openpyxl.styles.fills import GradientFill
wb = load_workbook('E:\\study\\全栈数据分析\\0812\\excle/案例.xlsx')
sheet = wb['各省销量']

#设置单元格字体
sheet['A1'].font = Font(
        name = '黑体', #设置单元格字体
        size = 36, #设置字体字号
        bold = True, #加粗
        color = 'FFaa8844' #单元格文本颜色
        )

#填充渐变颜色（起始颜色，结束颜色）
sheet['B1'].fill = GradientFill(stop = ('00ffff','00ffff'))

wb.save('E:\\study\\全栈数据分析\\0812\\excle/案例.xlsx')
